import streamlit as st
import pandas as pd
import numpy as np
import os
from pathlib import Path
from datetime import datetime
import plotly.graph_objects as go
from plotly.subplots import make_subplots

# Imports para geração dos relatórios profissionalmente
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors

# Configuração da página
st.set_page_config(page_title="Thereza - Incerteza Analítica", layout="wide")

APP_DIR = Path(__file__).resolve().parent
DATA_DIR = APP_DIR / "thereza_dados"
DATA_DIR.mkdir(exist_ok=True)

# --- INICIALIZAÇÃO DE ESTADO (SESSION STATE) ---
if "reg_results" not in st.session_state:
    st.session_state.reg_results = None
if "df_params" not in st.session_state:
    st.session_state.df_params = pd.DataFrame()
if "dados_curva" not in st.session_state:
    st.session_state.dados_curva = pd.DataFrame()
if "df_vert" not in st.session_state:
    st.session_state.df_vert = pd.DataFrame()
if "resultados_report" not in st.session_state:
    st.session_state.resultados_report = []
if "mapa_incertezas" not in st.session_state:
    st.session_state.mapa_incertezas = {}
if "dados_cca_report" not in st.session_state:
    st.session_state.dados_cca_report = {}

# --- FUNÇÕES DE PERSISTÊNCIA ---
def caminho_dados(composto, tipo="params"):
    prefixo = "config" if tipo == "params" else "curva"
    nome_seguro = composto.lower().replace("-", "_").replace(" ", "_")
    return DATA_DIR / f"{prefixo}_{nome_seguro}.csv"

def carregar_dados(composto, tipo="params"):
    arquivo = caminho_dados(composto, tipo)
    if arquivo.exists(): 
        df = pd.read_csv(arquivo, encoding="utf-8-sig")
        if tipo == "params" and not {"Parametro", "Valor"}.issubset(df.columns):
            st.warning(f"Arquivo de parâmetros inválido: {arquivo.name}. Usando valores padrão.")
            return None
        if tipo == "params" and "0. LOQ do Método (mg/Kg)" not in df["Parametro"].values:
            novo_loq = pd.DataFrame({"Parametro": ["0. LOQ do Método (mg/Kg)"], "Valor": [0.02]})
            df = pd.concat([novo_loq, df], ignore_index=True)
        return df
    return None

def salvar_dados(df, composto, tipo="params"):
    arquivo = caminho_dados(composto, tipo)
    df.to_csv(arquivo, index=False, encoding="utf-8-sig")

# --- FUNÇÕES DE GERAÇÃO DE RELATÓRIO ---
def gerar_excel(lista_dados, analista, data_hora, loq_atual, dados_cca=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório de Incerteza"
    
    ws['A1'] = "Avaliação da Incerteza da Medição Analítica"
    ws['A1'].font = Font(name='Arial', size=16, bold=True)
    ws['A2'] = f"Analista: {analista}"
    ws['A2'].font = Font(name='Arial', size=11, italic=True)
    ws['A3'] = f"Data/Hora: {data_hora}"
    ws['A3'].font = Font(name='Arial', size=11, italic=True)
    
    titulos = ["Amostras", "Valor mg/Kg", "Valor +/- U mg/Kg", "U%"]
    inclui_cca = dados_cca and len(dados_cca) > 0
    if inclui_cca:
        titulos.extend(["LMR (mg/Kg)", "CCα Limiar (mg/Kg)", "Status de Conformidade"])

    for col_num, titulo in enumerate(titulos, 1):
        celula = ws.cell(row=5, column=col_num)
        celula.value = titulo
        celula.font = Font(name='Arial', size=11, bold=True)
        celula.alignment = Alignment(horizontal='center')
        
    linha_atual = 6
    for dados in lista_dados:
        am_nome = dados["Amostra"]
        ws.cell(row=linha_atual, column=1, value=am_nome).alignment = Alignment(horizontal='left')
        ws.cell(row=linha_atual, column=2, value=dados["Valor_Exibicao"]).alignment = Alignment(horizontal='right')
        ws.cell(row=linha_atual, column=3, value=dados["Valor_U"]).alignment = Alignment(horizontal='center')
        ws.cell(row=linha_atual, column=4, value=dados["U_porcentagem_Exibicao"]).alignment = Alignment(horizontal='right')
        
        if inclui_cca:
            if am_nome in dados_cca:
                info_cca = dados_cca[am_nome]
                ws.cell(row=linha_atual, column=5, value=info_cca["lmr"]).alignment = Alignment(horizontal='right')
                ws.cell(row=linha_atual, column=6, value=round(info_cca["cca_limite"], 4)).alignment = Alignment(horizontal='right')
                ws.cell(row=linha_atual, column=7, value=info_cca["status_limpo"]).alignment = Alignment(horizontal='left')
            else:
                for c_vazia in range(5, 8):
                    ws.cell(row=linha_atual, column=c_vazia, value="-").alignment = Alignment(horizontal='center')

        for c in range(1, len(titulos) + 1):
            ws.cell(row=linha_atual, column=c).font = Font(name='Arial', size=11)
        linha_atual += 1
        
    linha_atual += 2
    ws.cell(row=linha_atual, column=1, value="Parecer Técnico e Diagnóstico Metrológico:").font = Font(name='Arial', size=12, bold=True)
    linha_atual += 1
    
    for dados in lista_dados:
        ws.cell(row=linha_atual, column=1, value=f"• {dados['Amostra']}: {dados['Parecer']}").font = Font(name='Arial', size=10)
        linha_atual += 1
        
    if inclui_cca:
        linha_atual += 1
        ws.cell(row=linha_atual, column=1, value="Parecer de Avaliação Legal de Risco (CCα):").font = Font(name='Arial', size=12, bold=True)
        linha_atual += 1
        for am_nome, info_cca in dados_cca.items():
            ws.cell(row=linha_atual, column=1, value=f"• {am_nome}: {info_cca['status_limpo']} (Abordagem: {info_cca['abordagem']})").font = Font(name='Arial', size=10)
            linha_atual += 1
        
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max(max_len + 3, 14)
        
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()

def gerar_pdf(lista_dados, analista, data_hora, dados_cca=None):
    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=letter, rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40)
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle('TitleStyle', parent=styles['Heading1'], fontName='Helvetica-Bold', fontSize=18, spaceAfter=12)
    meta_style = ParagraphStyle('MetaStyle', parent=styles['Normal'], fontName='Helvetica-Oblique', fontSize=11, spaceAfter=4)
    table_text = ParagraphStyle('TableText', parent=styles['Normal'], fontName='Helvetica', fontSize=10, alignment=0)
    table_text_center = ParagraphStyle('TableTextCenter', parent=styles['Normal'], fontName='Helvetica', fontSize=10, alignment=1)
    table_header = ParagraphStyle('TableHeader', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=10, alignment=1)
    section_heading = ParagraphStyle('SecHeading', parent=styles['Heading2'], fontName='Helvetica-Bold', fontSize=12, spaceBefore=20, spaceAfter=8)
    parecer_style = ParagraphStyle('ParecerStyle', parent=styles['Normal'], fontName='Helvetica', fontSize=10, spaceAfter=6, leftIndent=10)
    
    elementos = []
    
    elementos.append(Paragraph("Avaliação da Incerteza da Medição Analítica", title_style))
    elementos.append(Paragraph(f"Analista: {analista}", meta_style))
    elementos.append(Paragraph(f"Data/Hora: {data_hora}", meta_style))
    elementos.append(Spacer(1, 15))
    
    inclui_cca = dados_cca and len(dados_cca) > 0
    
    if inclui_cca:
        dados_tabela = [[
            Paragraph("Amostras", table_header),
            Paragraph("Valor nominal", table_header),
            Paragraph("Valor +/- U", table_header),
            Paragraph("U%", table_header),
            Paragraph("LMR", table_header),
            Paragraph("Limiar CCα", table_header)
        ]]
        col_widths = [110, 80, 110, 60, 60, 90]
        
        for row in lista_dados:
            am_nome = row["Amostra"]
            lmr_txt, cca_txt = "-", "-"
            if am_nome in dados_cca:
                lmr_txt = f"{dados_cca[am_nome]['lmr']:.3f}"
                cca_txt = f"{dados_cca[am_nome]['cca_limite']:.4f}"
                
            dados_tabela.append([
                Paragraph(am_nome, table_text),
                Paragraph(row["Valor_Exibicao"], table_text_center),
                Paragraph(row["Valor_U"], table_text_center),
                Paragraph(row["U_porcentagem_Exibicao"], table_text_center),
                Paragraph(lmr_txt, table_text_center),
                Paragraph(cca_txt, table_text_center)
            ])
    else:
        dados_tabela = [[
            Paragraph("Amostras", table_header),
            Paragraph("Valor mg/Kg", table_header),
            Paragraph("Valor +/- U mg/Kg", table_header),
            Paragraph("U%", table_header)
        ]]
        col_widths = [150, 110, 160, 90]
        
        for row in lista_dados:
            dados_tabela.append([
                Paragraph(row["Amostra"], table_text),
                Paragraph(row["Valor_Exibicao"], table_text_center),
                Paragraph(row["Valor_U"], table_text_center),
                Paragraph(row["U_porcentagem_Exibicao"], table_text_center)
            ])
    
    t = Table(dados_tabela, colWidths=col_widths)
    t.setStyle(TableStyle([
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('LINEBELOW', (0,0), (-1,0), 1, colors.HexColor("#333333")), 
    ]))
    elementos.append(t)
    
    elementos.append(Paragraph("Parecer Técnico e Diagnóstico Metrológico", section_heading))
    for row in lista_dados:
        texto_parecer = f"<b>{row['Amostra']}:</b> {row['Parecer']}"
        elementos.append(Paragraph(texto_parecer, parecer_style))
        
    if inclui_cca:
        elementos.append(Paragraph("Parecer de Avaliação Legal de Risco (CCα)", section_heading))
        for am_nome, info_cca in dados_cca.items():
            texto_cca = f"<b>{am_nome}:</b> {info_cca['status_limpo']} <i>(Abordagem normativa: {info_cca['abordagem']} / Risco α: {info_cca['alfa']}%).</i>"
            elementos.append(Paragraph(texto_cca, parecer_style))
        
    doc.build(elementos)
    return output.getvalue()


# --- SIDEBAR COM OS COMPOSTOS ---
with st.sidebar:
    if os.path.exists("logo_thereza.png"):
        st.image("logo_thereza.png", width="stretch")
    elif os.path.exists("logo_thereza.png.png"):
        st.image("logo_thereza.png.png", width="stretch")
    else:
        st.markdown("<h2 style='color:#ffffff; text-align:center; font-weight:bold;'>THEREZA</h2>", unsafe_allow_html=True)
        st.caption("Cálculo de Incerteza de Medição")
        
    st.markdown("---")
    st.header("⚙️ Configuração")
    
    metodo = st.radio("Método:", ["Glifosato", "Multirresíduos"], key="metodo_sidebar")
    
    if metodo == "Glifosato":
        opcoes_composto = ["Glifosato", "Glufosinato"]
    else:
        opcoes_composto = [
            "Abamectina", "Atrazina", "Azoxistrobina", "Clorantraniliprole", 
            "Clorpirifós", "Clorpirifós-metílico", "Clotianidina", "Ciproconazol", 
            "Epoxiconazol", "Flupiradifurona", "Flutriafol", "Fluxapiroxade", 
            "Imidacloprido", "Pencicurona", "Piraclostrobina", "Tiametoxam", "Triadimenol"
        ]
    
    composto = st.selectbox("Composto:", options=opcoes_composto, key="composto_selecionado")
    
    st.markdown("---")
    st.header("👤 Operação")
    
    analista_selecionado = st.selectbox("Selecione o Analista:", ["Yasmim", "Núbia"], key="analista_input")


# --- ESTILIZAÇÃO CSS ---
st.markdown("""
    <style>
    [data-testid="stSidebar"] { background-color: #0a2540 !important; }
    [data-testid="stSidebar"] h1, [data-testid="stSidebar"] h2, [data-testid="stSidebar"] h3, 
    [data-testid="stSidebar"] p, [data-testid="stSidebar"] span, [data-testid="stSidebar"] label,
    [data-testid="stSidebar"] caption { color: #ffffff !important; }
    [data-testid="stSidebar"] div[data-baseweb="select"] * { color: #111111 !important; }
    [data-testid="stSidebar"] input { color: #111111 !important; }
    [data-testid="stSidebar"] div[role="radiogroup"] label span { color: #ffffff !important; }
    
    .metric-container { background-color: #ffffff; padding: 10px; border-radius: 8px; border: 1px solid #e0e0e0; text-align: center; box-shadow: 0 2px 4px rgba(0,0,0,0.05); min-width: 120px; }
    .metric-label { color: #333333; font-size: 11px; font-weight: 700; margin-bottom: 2px; }
    .metric-value { color: #0a2540; font-size: 14px; font-family: 'Courier New', monospace; font-weight: bold; }
    .uc-highlight { border-top: 4px solid #00d4ff; }
    .diag-box { padding: 15px; border-radius: 5px; margin-top: 15px; border-left: 5px solid; font-weight: normal; font-size: 13px; line-height: 1.5; }
    .diag-good { background-color: #e8f5e9; border-left-color: #2e7d32; color: #1b5e20; }
    .diag-warning { background-color: #fff3e0; border-left-color: #ef6c00; color: #e65100; }
    .param-label { font-size: 12px; color: #555555; }
    .param-value { font-size: 18px; font-weight: bold; color: #0a2540; }
    .res-box { background-color: #f4f6f8; padding: 20px; border-radius: 10px; border: 1px solid #dee2e6; margin-bottom: 25px; border-left: 6px solid #0a2540; }
    .sample-title { font-size: 1.2rem; font-weight: bold; color: #0a2540; margin: 0; }
    .main-result { font-size: 1.8rem; color: #0a2540; font-weight: bold; text-align: center; margin: 0; }
    .stExpander { border: none !important; box-shadow: none !important; }
    </style>
    """, unsafe_allow_html=True)


# --- INTERCEPTAÇÃO E ATRIBUIÇÃO DE VALORES PADRÃO ---
if "composto_atual" not in st.session_state or st.session_state.composto_atual != composto:
    st.session_state.composto_atual = combo_composto = composto
    
    df_p = carregar_dados(composto, "params")
    
    if df_p is None:
        valores_padrao = {
            "0. LOQ do Método (mg/Kg)": 0.02,
            "1. Balança - U(tara)": 0.0001,
            "1. Balança - U(pesagem)": 0.0001,
            "1. Balança - k calibração": 2.0,
            "2. Padrão - Pureza (%)": 99.5,
            "2. Padrão - U(padrão) certified": 0.02,
            "3. Vse - Volume solução extratora (mL)": 10.0,
            "3. Vse - U(pipeta)": 0.02,
            "4. Repetibilidade - RSD (%)": 5.4,
            "4. Repetibilidade - Graus Liberdade": 26.0,
            "5. Vol IS (mL)": 0.05,
            "5. Vol IS - U(pipetador)": 0.001,
            "6. Vol Amostra (mL)": 0.45,
            "6. Vol Amostra - U(pipetador)": 0.005,
        }
        
        # Adiciona os parâmetros específicos do solvente se for Multirresíduos
        if metodo == "Multirresíduos":
            valores_padrao["6. Vol Amostra (mL)"] = 0.50
            valores_padrao["7. Vol Solvente (mL)"] = 0.50
            valores_padrao["7. Vol Solvente - U(pipetador)"] = 0.005
            
        if composto == "Clorpirifós":
            valores_padrao["4. Repetibilidade - RSD (%)"] = 4.8
        elif composto == "Glifosato":
            valores_padrao["0. LOQ do Método (mg/Kg)"] = 0.05
            valores_padrao["3. Vse - Volume solução extratora (mL)"] = 5.0
            
        lista_parametros = list(valores_padrao.keys())
        lista_valores = list(valores_padrao.values())
        df_p = pd.DataFrame({"Parametro": lista_parametros, "Valor": lista_valores})

    if metodo == "Multirresíduos":
        df_p["Parametro"] = df_p["Parametro"].replace({
            "5. Vol Solvente (mL)": "7. Vol Solvente (mL)",
            "5. Vol Solvente - U(pipetador)": "7. Vol Solvente - U(pipetador)",
        })
        protocolo_multirresiduos = {
            "6. Vol Amostra (mL)": 0.50,
            "7. Vol Solvente (mL)": 0.50,
        }
        for param_nome, param_valor in protocolo_multirresiduos.items():
            if param_nome in df_p["Parametro"].values:
                df_p.loc[df_p["Parametro"] == param_nome, "Valor"] = param_valor
            else:
                df_p = pd.concat(
                    [df_p, pd.DataFrame({"Parametro": [param_nome], "Valor": [param_valor]})],
                    ignore_index=True
                )

    st.session_state.df_params = df_p
    
    df_c = carregar_dados(composto, "curva")
    st.session_state.dados_curva = df_c if df_c is not None else pd.DataFrame({"Conc (mg/L)": [0.0]*6, "Resposta": [0.0]*6})
    st.session_state.reg_results = None
    st.session_state.resultados_report = []
    st.session_state.mapa_incertezas = {}
    st.session_state.dados_cca_report = {}

# --- MAPEAMENTO DO LOQ E PARÂMETROS DE ENTRADA ---
p = dict(zip(st.session_state.df_params["Parametro"], st.session_state.df_params["Valor"]))

LOQ_METODO = p.get("0. LOQ do Método (mg/Kg)", 0.02)

k_calib = p.get("1. Balança - k calibração", 1)
if k_calib == 0 or k_calib is None: k_calib = 1

u_massa = np.sqrt((p.get("1. Balança - U(tara)", 0) / k_calib)**2 + (p.get("1. Balança - U(pesagem)", 0) / k_calib)**2)
u_pureza = (p.get("2. Padrão - U(padrão) certified", 0) / 100) / np.sqrt(3)
u_vse = p.get("3. Vse - U(pipeta)", 0) / np.sqrt(3)
u_rep = p.get("4. Repetibilidade - RSD (%)", 0) / 100
u_vis = p.get("5. Vol IS - U(pipetador)", 0) / np.sqrt(3)
u_vam = p.get("6. Vol Amostra - U(pipetador)", 0) / np.sqrt(3)

# Lógica condicional metrológica para o cálculo do Volume Final e suas incertezas
v_se_val = p.get("3. Vse - Volume solução extratora (mL)", 0)
v_is_val = p.get("5. Vol IS (mL)", 0)
v_am_val = p.get("6. Vol Amostra (mL)", 0)

if metodo == "Multirresíduos":
    v_solvente_val = p.get("7. Vol Solvente (mL)", p.get("5. Vol Solvente (mL)", 0.50))
    u_vsolvente = p.get("7. Vol Solvente - U(pipetador)", p.get("5. Vol Solvente - U(pipetador)", 0.005)) / np.sqrt(3)
    
    v_final_val = v_am_val + v_solvente_val + v_is_val
    u_vf = np.sqrt(u_vam**2 + u_vsolvente**2 + u_vis**2)
    texto_processo = f"**Processo (Multirresíduos):** Vse = {v_se_val:.2f} mL | Vol Extrato = {v_am_val:.2f} mL | Vol Solvente = {v_solvente_val:.2f} mL | Vol IS = {v_is_val:.2f} mL | Volume Final = {v_final_val:.2f} mL"
else:
    v_solvente_val = 0.0
    u_vsolvente = 0.0
    v_final_val = 0.50  # Valor fixo do balão/extrato original do Glifosato
    u_vf = np.sqrt(u_vis**2 + u_vam**2)
    texto_processo = f"**Processo (Glifosato):** Vse = {v_se_val:.2f} mL | Volume Final Amostra = {(v_is_val + v_am_val):.2f} mL"

st.subheader(f"📊 Parâmetros de Incerteza de Entrada: {composto}")
all_u = [("LOQ Vigente", LOQ_METODO), ("u Massa", u_massa), ("u Padrão", u_pureza), ("u Vse", u_vse), ("u Vf", u_vf), ("u V(Amostra)", u_vam), ("u Repetib.", u_rep)]

cols_u = st.columns(len(all_u))
for i, (label, val) in enumerate(all_u):
    with cols_u[i]:
        str_val = f"{val:.3f}" if label == "LOQ Vigente" else f"{val:.4e}"
        st.markdown(f'<div class="metric-container uc-highlight"><div class="metric-label">{label}</div><div class="metric-value">{str_val}</div></div>', unsafe_allow_html=True)

st.info(texto_processo)

with st.expander("📝 Editar Valores de Entrada (Estratégia 1: Armazenamento Local)"):
    uploaded_config = st.file_uploader("📂 Carregar arquivo de configuração personalizado (CSV)", type=["csv"], key="uploader_config_local")
    if uploaded_config is not None:
        try:
            df_uploaded = pd.read_csv(uploaded_config)
            if "Parametro" in df_uploaded.columns and "Valor" in df_uploaded.columns:
                st.session_state.df_params = df_uploaded
                st.success("📝 Configuração local carregada com sucesso!")
            else:
                st.error("⚠️ O arquivo CSV enviado não possui o formato esperado (colunas 'Parametro' e 'Valor').")
        except Exception as e:
            st.error(f"Erro ao ler o arquivo: {e}")

    df_ed = st.data_editor(st.session_state.df_params, hide_index=True, width='stretch')
    
    c_btn1, c_btn2 = st.columns(2)
    with c_btn1:
        if st.button("💾 Salvar Alterações na Memória"):
            st.session_state.df_params = df_ed
            salvar_dados(df_ed, composto, "params")
            st.success(f"Parâmetros updated com sucesso para {composto}!")
            st.rerun()
            
    with c_btn2:
        csv_buffer = io.StringIO()
        df_ed.to_csv(csv_buffer, index=False)
        st.download_button(
            label="📥 Baixar esta Configuração (CSV Local)",
            data=csv_buffer.getvalue(),
            file_name=f"config_{composto.lower().replace('-', '_').replace(' ', '_')}_meu_lab.csv",
            mime="text/csv",
            width='stretch'
        )

st.markdown("---")

# --- SEÇÃO 2: ESTATÍSTICA DA CURVA ---
st.subheader("📈 Estatística da Curva")

cc1, cc2 = st.columns(2)
with cc1:
    n_niveis = st.number_input("Número de Níveis da Curva:", min_value=3, max_value=15, value=int(len(st.session_state.dados_curva)))
with cc2:
    tipo_curva = st.selectbox("Tipo de Curva de Calibração:", ["Linear", "Linear Ponderada 1/x", "Linear Ponderada 1/x2"])

if len(st.session_state.dados_curva) != n_niveis:
    nova_curva = pd.DataFrame({"Conc (mg/L)": [0.0]*n_niveis, "Resposta": [0.0]*n_niveis})
    for col in ["Conc (mg/L)", "Resposta"]:
        min_len = min(len(st.session_state.dados_curva), n_niveis)
        nova_curva.iloc[:min_len, nova_curva.columns.get_loc(col)] = st.session_state.dados_curva.iloc[:min_len, nova_curva.columns.get_loc(col)]
    st.session_state.dados_curva = nova_curva

df_curva = st.data_editor(st.session_state.dados_curva, hide_index=True, width='stretch')
st.session_state.dados_curva = df_curva

if st.button("📊 Calcular e Fixar Curva"):
    df_curva_calc = df_curva.copy()
    df_curva_calc["Conc (mg/L)"] = pd.to_numeric(df_curva_calc["Conc (mg/L)"], errors="coerce")
    df_curva_calc["Resposta"] = pd.to_numeric(df_curva_calc["Resposta"], errors="coerce")
    df_curva_calc = df_curva_calc.replace([np.inf, -np.inf], np.nan).dropna()
    x, y = df_curva_calc["Conc (mg/L)"].values, df_curva_calc["Resposta"].values
    n = len(x)

    erro_curva = None
    if n < 3:
        erro_curva = "Informe pelo menos 3 pontos válidos na curva."
    elif np.any(x < 0) or np.any(y < 0):
        erro_curva = "A curva não deve conter concentração ou resposta negativa."
    elif len(np.unique(x)) < 2:
        erro_curva = "Informe ao menos duas concentrações diferentes na curva."

    if erro_curva:
        st.error(f"⚠️ {erro_curva}")
        st.session_state.reg_results = None
    else:
        if tipo_curva == "Linear Ponderada 1/x":
            w = np.array([1.0 / xi if xi > 0 else 1.0 / 1e-5 for xi in x])
        elif tipo_curva == "Linear Ponderada 1/x2":
            w = np.array([1.0 / (xi**2) if xi > 0 else 1.0 / 1e-10 for xi in x])
        else: 
            w = np.ones(n)
        
        sw = np.sum(w)
        swx = np.sum(w * x)
        swy = np.sum(w * y)
        swx2 = np.sum(w * (x**2))
        swxy = np.sum(w * x * y)
    
        delta = sw * swx2 - (swx**2)
        if np.isclose(delta, 0):
            st.error("⚠️ Não foi possível calcular a curva: os pontos informados não geram regressão válida.")
            st.session_state.reg_results = None
        else:
            a = (sw * swxy - swx * swy) / delta
            b = (swx2 * swy - swx * swxy) / delta
    
            y_pred = a * x + b
            r = np.corrcoef(x, y)[0,1]
            r2 = r**2
    
            epadyx = np.sqrt(np.sum(w * ((y - y_pred)**2)) / (n - 2))
            ss_xx = np.sum(w * ((x - (swx/sw))**2))
    
            sa = epadyx / np.sqrt(ss_xx)
            sb = epadyx * np.sqrt(np.sum(w * (x**2)) / (sw * ss_xx))
    
            st.session_state.reg_results = {
                "a":a, "b":b, "sa":sa, "sb":sb, "r":r, "r2":r2, 
                "epadyx":epadyx, "gl": n-2, "x":x, "y":y, "y_pred":y_pred, "tipo": tipo_curva
            }
            salvar_dados(df_curva_calc, composto, "curva")

if st.session_state.reg_results:
    res = st.session_state.reg_results
    st.markdown("### 📌 Parâmetros da Curva")
    
    row1 = st.columns(4)
    row1[0].markdown(f"<span class='param-label'>Slope (a)</span><br><span class='param-value'>{res['a']:.4e}</span>", unsafe_allow_html=True)
    row1[1].markdown(f"<span class='param-label'>Intercept (b)</span><br><span class='param-value'>{res['b']:.4e}</span>", unsafe_allow_html=True)
    row1[2].markdown(f"<span class='param-label'>s_a</span><br><span class='param-value'>{res['sa']:.4e}</span>", unsafe_allow_html=True)
    row1[3].markdown(f"<span class='param-label'>s_b</span><br><span class='param-value'>{res['sb']:.4e}</span>", unsafe_allow_html=True)

    row2 = st.columns(4)
    row2[0].markdown(f"<span class='param-label'>r</span><br><span class='param-value'>{res['r']:.5f}</span>", unsafe_allow_html=True)
    row2[1].markdown(f"<span class='param-label'>R²</span><br><span class='param-value'>{res['r2']:.5f}</span>", unsafe_allow_html=True)
    row2[2].markdown(f"<span class='param-label'>EPADYX (s_y/x)</span><br><span class='param-value'>{res['epadyx']:.4e}</span>", unsafe_allow_html=True)
    row2[3].markdown(f"<span class='param-label'>GL</span><br><span class='param-value'>{res['gl']}</span>", unsafe_allow_html=True)

    st.markdown("### 📋 Interpretação e Diagnóstico da Curva")
    if res['r2'] >= 0.990:
        st.markdown(f"""
        <div class="diag-box diag-good">
            <strong>✔️ Modelo Linear Adequado:</strong> O coeficiente de determinação (R² = {res['r2']:.5f}) atende aos critérios de aceitação regulamentares. A sensibilidade do método (Slope) está consolidada em {res['a']:.4e} e o erro residual da curva (s_y/x) apresenta baixa magnitude ({res['epadyx']:.4e}).
        </div>
        """, unsafe_allow_html=True)
    else:
        st.markdown(f"""
        <div class="diag-box diag-warning">
            <strong>⚠️ Atenção - Necessita Ajuste ou Troca de Curva:</strong> O coeficiente de determinação (R² = {res['r2']:.5f}) encontra-se abaixo do limiar estatístico ideal de 0,990.
        </div>
        """, unsafe_allow_html=True)

    fig = make_subplots(rows=1, cols=3, subplot_titles=("Curva", "Resíduos Y", "Resíduos X"))
    fig.add_trace(go.Scatter(x=res['x'], y=res['y'], mode='markers', name="Dados", marker=dict(color='#0a2540')), row=1, col=1)
    fig.add_trace(go.Scatter(x=res['x'], y=res['y_pred'], mode='lines', name="Ajuste", line=dict(color='#00d4ff')), row=1, col=1)
    
    fig.add_trace(go.Scatter(x=res['x'], y=res['y']-res['y_pred'], mode='markers', marker=dict(color='#ef6c00')), row=1, col=2)
    fig.add_shape(type="line", x0=min(res['x']), y0=0, x1=max(res['x']), y1=0, line=dict(dash="dash", color="black"), row=1, col=2)
    
    res_x = ((res['y'] - res['b'])/res['a']) - res['x']
    fig.add_trace(go.Scatter(x=res['x'], y=res_x, mode='markers', marker=dict(color='green')), row=1, col=3)
    fig.add_shape(type="line", x0=min(res['x']), y0=0, x1=max(res['x']), y1=0, line=dict(dash="dash", color="black"), row=1, col=3)
    
    fig.update_layout(height=350, template="plotly_white", showlegend=False)
    st.plotly_chart(fig, width='stretch')

st.markdown("---")

# --- SEÇÃO 3: MEDIÇÃO DE AMOSTRAS ---
st.header("🧪 Medição de Amostras")
ca1, ca2 = st.columns(2)
n_am, n_rep = ca1.number_input("Qtd Amostras:", 1, 10, 1), ca2.number_input("Replicatas:", 1, 10, 2)

total_linhas = n_am * n_rep
massa_padrao_amostra = 5.00 if metodo == "Multirresíduos" else 2.01
if "metodo_amostras_atual" not in st.session_state or st.session_state.metodo_amostras_atual != metodo:
    st.session_state.metodo_amostras_atual = metodo
    st.session_state.df_vert = pd.DataFrame()

if len(st.session_state.df_vert) != total_linhas:
    nomes_amostras = []
    for i in range(n_am):
        for j in range(n_rep):
            nomes_amostras.append(f"Amostra_{i+1}")
            
    st.session_state.df_vert = pd.DataFrame({
        "Amostra": nomes_amostras, 
        "Resposta": [0.0] * total_linhas, 
        "Massa (g)": [massa_padrao_amostra] * total_linhas
    })

df_med = st.data_editor(st.session_state.df_vert, hide_index=True, width='stretch')
st.session_state.df_vert = df_med

if st.button("🚀 Gerar Avaliação de Incerteza"):
    if st.session_state.get("reg_results"):
        res = st.session_state.reg_results
        
        df_processado = df_med.copy()
        df_processado["Resposta"] = pd.to_numeric(df_processado["Resposta"], errors="coerce")
        df_processado["Massa (g)"] = pd.to_numeric(df_processado["Massa (g)"], errors="coerce")
        df_processado = df_processado.replace([np.inf, -np.inf], np.nan).dropna(subset=["Amostra", "Resposta", "Massa (g)"])
        df_processado = df_processado[df_processado["Massa (g)"] > 0]

        if df_processado.empty:
            st.error("⚠️ Informe ao menos uma amostra válida, com resposta numérica e massa maior que zero.")
            st.stop()
        if np.isclose(res["a"], 0):
            st.error("⚠️ A inclinação da curva está zerada ou muito próxima de zero. Recalcule a curva antes da avaliação.")
            st.stop()
        if v_se_val <= 0 or v_am_val <= 0 or v_final_val <= 0:
            st.error("⚠️ Verifique os volumes de entrada: Vse, V(amostra) e volume final devem ser maiores que zero.")
            st.stop()

        df_processado['Amostra_Grupo'] = df_processado['Amostra'].apply(lambda x: str(x).split('_')[0] if '_' in str(x) else str(x))
        
        df_agrupado = df_processado.groupby('Amostra_Grupo').agg(
            Resposta_Media=('Resposta', 'mean'),
            Resposta_Desvio=('Resposta', 'std'),
            Massa_Media=('Massa (g)', 'mean'),
            Contagem_Replicas=('Resposta', 'count')
        ).reset_index()
        
        lista_temporaria_report = []
        mapa_temporario_u = {}
        
        for _, row in df_agrupado.iterrows():
            nome_amostra = row['Amostra_Grupo']
            y_am = row['Resposta_Media']
            m_am = row['Massa_Media']
            n_reps_atual = int(row['Contagem_Replicas'])
            
            if n_reps_atual > 1 and not pd.isna(row['Resposta_Desvio']):
                u_resposta_dinamica = row['Resposta_Desvio']
                gl_resposta_dinamica = n_reps_atual - 1
            else:
                u_resposta_dinamica = res['epadyx']
                gl_resposta_dinamica = 1
            
            # Ajuste do fator de conversão de diluição baseado no método selecionado
            pureza_fracao = p.get("2. Padrão - Pureza (%)", 99.4) / 100
            fator_repetitividade = 1.0
            conc_base_sem_pureza = None
            incerteza_pureza_tabela = p.get("2. Padrão - U(padrão) certified", 0.0173)
            ci_pureza = 0
            ci_repetitividade = 0

            if metodo == "Multirresíduos":
                fator_diluicao = (v_se_val * v_final_val) / (m_am * v_am_val)
                conc_base_sem_pureza = ((y_am - res['b']) / res['a']) * fator_diluicao
                conc = conc_base_sem_pureza * fator_repetitividade * pureza_fracao
                f_comum = fator_diluicao * fator_repetitividade * pureza_fracao
                incerteza_pureza_tabela = u_pureza
                ci_pureza = conc_base_sem_pureza * fator_repetitividade
                ci_repetitividade = conc / fator_repetitividade if fator_repetitividade != 0 else 0
            else:
                f_comum = v_se_val / (m_am * (v_am_val / (v_is_val + v_am_val)))
                conc = (((y_am - res['b']) / res['a']) * v_se_val) / (m_am * (v_am_val / (v_is_val + v_am_val)))
                ci_pureza = conc / (p.get("2. Padrão - Pureza (%)", 99.4) * np.sqrt(3)) if p.get("2. Padrão - Pureza (%)", 99.4) > 0 else 0
                ci_repetitividade = conc

            gl_repetibilidade = int(p.get("4. Repetibilidade - Graus Liberdade", 26))
            
            # Quadro de incerteza preservando o cálculo travado
            linhas_tabela = [
                {"Grandeza": "Resposta", "Valor da grandeza": y_am, "Valor da incerteza": u_resposta_dinamica, "Unidade": "unidades área", "Distribuição": "Normal", "Graus de liberdade": str(gl_resposta_dinamica), "ci": (1 / res['a']) * f_comum},
                {"Grandeza": "INTERCEPTO", "Valor da grandeza": res['b'], "Valor da incerteza": res['sb'], "Unidade": "unidades área", "Distribuição": "Normal", "Graus de liberdade": str(res['gl']), "ci": (-1 / res['a']) * f_comum},
                {"Grandeza": "SLOPE", "Valor da grandeza": res['a'], "Valor da incerteza": res['sa'], "Unidade": "unidades área", "Distribuição": "Normal", "Graus de liberdade": str(res['gl']), "ci": -((y_am - res['b']) / (res['a'] ** 2)) * f_comum},
                {"Grandeza": "Massa amostra", "Valor da grandeza": m_am, "Valor da incerteza": u_massa, "Unidade": "g", "Distribuição": "Normal", "Graus de liberdade": "infinito", "ci": -conc / m_am if m_am > 0 else 0},
                {"Grandeza": "Pureza padrão", "Valor da grandeza": p.get("2. Padrão - Pureza (%)", 99.4), "Valor da incerteza": incerteza_pureza_tabela, "Unidade": "%", "Distribuição": "Retangular", "Graus de liberdade": "infinito", "ci": ci_pureza},
                {"Grandeza": "V(se)", "Valor da grandeza": v_se_val, "Valor da incerteza": u_vse, "Unidade": "mL", "Distribuição": "Retangular", "Graus de liberdade": "infinito", "ci": conc / v_se_val if v_se_val > 0 else 0},
                {"Grandeza": "V(f)", "Valor da grandeza": v_final_val, "Valor da incerteza": u_vf, "Unidade": "mL", "Distribuição": "Retangular", "Graus de liberdade": "infinito", "ci": conc / v_final_val if v_final_val > 0 else 0},
                {"Grandeza": "V(amostra)", "Valor da grandeza": v_am_val, "Valor da incerteza": u_vam, "Unidade": "mL", "Distribuição": "Retangular", "Graus de liberdade": "infinito", "ci": -conc / v_am_val if v_am_val > 0 else 0},
                {"Grandeza": "Repetitividade", "Valor da grandeza": fator_repetitividade, "Valor da incerteza": u_rep, "Unidade": "-", "Distribuição": "Normal", "Graus de liberdade": str(gl_repetibilidade), "ci": ci_repetitividade}
            ]
            
            df_detalhado = pd.DataFrame(linhas_tabela)
            
            def converter_gl(valor):
                if str(valor).strip().lower() == 'infinito': return np.inf
                return float(valor)
            
            df_detalhado["gl_num"] = df_detalhado["Graus de liberdade"].apply(converter_gl)
            df_detalhado["Incerteza padrão, u(yi) = C_i*u_i"] = df_detalhado["ci"] * df_detalhado["Valor da incerteza"]
            df_detalhado["Incerteza, u(y_i)^2"] = df_detalhado["Incerteza padrão, u(yi) = C_i*u_i"] ** 2
            
            soma_variancias = df_detalhado["Incerteza, u(y_i)^2"].sum()
            uc_calc = np.sqrt(soma_variancias)
            
            termo_ws = 0.0
            for _, r_tabela in df_detalhado.iterrows():
                u_quarta = r_tabela["Incerteza, u(y_i)^2"] ** 2
                gl_i = r_tabela["gl_num"]
                if gl_i != np.inf and gl_i > 0:
                    termo_ws += u_quarta / gl_i
                    
            if termo_ws > 0:
                v_eff = (uc_calc ** 4) / termo_ws
                v_eff_trunc = int(round(v_eff))
            else:
                v_eff_trunc = 1000

            if v_eff_trunc >= 1000:
                k_dinamico = 2.000
            else:
                k_dinamico = 2.000 + (2.362 / v_eff_trunc) + (4.453 / (v_eff_trunc ** 2))
                k_dinamico = round(k_dinamico, 3)
            
            if soma_variancias <= 0:
                st.error(f"⚠️ Não foi possível calcular a contribuição de incerteza para {nome_amostra}. Verifique os dados informados.")
                continue

            df_detalhado["Contribuição_Num"] = (df_detalhado["Incerteza, u(y_i)^2"] / soma_variancias * 100)
            df_detalhado["Contribuição"] = df_detalhado["Contribuição_Num"].apply(lambda x: f"{x:.2f}%")
            
            U_exp = uc_calc * k_dinamico
            U_per = (U_exp / conc * 100) if conc > 0 else 0

            idx_max = df_detalhado["Contribuição_Num"].idxmax()
            componente_max = df_detalhado.loc[idx_max, "Grandeza"]
            
            if componente_max == "Repetitividade":
                texto_melhoria = "A repetitividade histórica é o principal factor no balanço de incerteza."
            elif componente_max in ["Resposta", "INTERCEPTO", "SLOPE"]:
                texto_melhoria = "A variabilidade instrumental domina o erro. Recomenda-se aumentar as replicatas."
            else:
                texto_melhoria = "Os componentes gravimétricos/volumétricos dominam a incerteza. Priorize vidrarias Classe A."

            if conc < LOQ_METODO:
                valor_exibicao_tela = f"< {LOQ_METODO:.3f} mg/Kg"
                valor_nominal_pdf = "-"
                valor_u_pdf = f"< {LOQ_METODO:.3f}"
                u_porcentagem_tela = "-"
                parecer_final = f"Resultado abaixo do Limite de Quantificação (LOQ) dinâmico do composto ({LOQ_METODO:.3f} mg/Kg)."
            else:
                valor_exibicao_tela = f"{conc:.3f} ± {U_exp:.3f} mg/Kg"
                valor_nominal_pdf = f"{conc:.3f}"
                valor_u_pdf = f"{conc:.3f} ± {U_exp:.3f}"
                u_porcentagem_tela = f"{U_per:.1f}%"
                if conc > df_curva["Conc (mg/L)"].max():
                    parecer_final = f"Concentração acima do topo calibrado da curva. Risco de saturação. {texto_melhoria}"
                else:
                    parecer_final = f"Resultado quantificado com sucesso dentro do intervalo linear. {texto_melhoria}"

            df_visible = df_detalhado[[
                "Grandeza", "Valor da grandeza", "Valor da incerteza", "Unidade", 
                "Distribuição", "Graus de liberdade", "ci", 
                "Incerteza padrão, u(yi) = C_i*u_i", "Incerteza, u(y_i)^2", "Contribuição"
            ]].rename(columns={"ci": "Coeficiente de sensibilidade, C_i"})

            mapa_temporario_u[nome_amostra] = {
                "conc": conc, 
                "uc": uc_calc, 
                "valor_exibicao_tela": valor_exibicao_tela,
                "u_porcentagem_tela": u_porcentagem_tela,
                "parecer_final": parecer_final,
                "df_visible": df_visible,
                "v_eff_trunc": v_eff_trunc,
                "k_dinamico": k_dinamico,
                "u_resposta_dinamica": u_resposta_dinamica,
                "n_reps_atual": n_reps_atual
            }

            lista_temporaria_report.append({
                "Amostra": nome_amostra,
                "Valor": conc,
                "Valor_Exibicao": valor_nominal_pdf,
                "Valor_U": valor_u_pdf,
                "U_porcentagem": U_per,
                "U_porcentagem_Exibicao": f"{U_per:.1f}%" if conc >= LOQ_METODO else "-",
                "Parecer": parecer_final
            })
            
        st.session_state.resultados_report = lista_temporaria_report
        st.session_state.mapa_incertezas = mapa_temporario_u
    else:
        st.error("⚠️ Calcule e fixe a curva analítica antes de rodar a avaliação de incerteza.")

# --- EXIBIÇÃO PERSISTENTE DOS QUADROS DE INCERTEZA ---
if st.session_state.mapa_incertezas:
    st.subheader("🏁 Avaliação da Incerteza de Medição Analítica")
    for nome_amostra, m_data in st.session_state.mapa_incertezas.items():
        with st.container():
            st.markdown('<div class="res-box">', unsafe_allow_html=True)
            r_col1, r_col2, r_col3 = st.columns([1.5, 4.0, 1.5])
            with r_col1: 
                st.markdown(f'<p class="sample-title">{nome_amostra}</p>', unsafe_allow_html=True)
            with r_col2: 
                st.markdown(f'<p class="main-result">{m_data["valor_exibicao_tela"]}</p>', unsafe_allow_html=True)
            with r_col3: 
                st.markdown(f'<p class="main-result" style="font-size:1.2rem; text-align:right;">{m_data["u_porcentagem_tela"]}</p>', unsafe_allow_html=True)
            
            st.markdown(f"<div style='font-size:13px; color:#0a2540; margin-top:5px; margin-bottom:10px;'><b>Nota de Parecer Técnico:</b> {m_data['parecer_final']}</div>", unsafe_allow_html=True)
            
            with st.expander("🔍 Detalhar o Balanço de Incerteza"):
                st.markdown(f"**📋 Quadro Completo de Incerteza - Identificação: {nome_amostra}**")
                st.dataframe(m_data["df_visible"], width='stretch', hide_index=True)
                
                st.markdown(f"""
                <p style='font-size:12px; color:#555; margin-top:10px;'>
                    <strong>Incerteza padrão combinada, u<sub>c</sub>:</strong> {m_data['uc']:.6f}<br>
                    <strong>Graus de liberdade efetivos (v<sub>eff</sub>):</strong> {m_data['v_eff_trunc']}<br>
                    <strong>Fator de abrangência k:</strong> {m_data['k_dinamico']:.3f} (calculado via Welch-Satterthwaite)<br>
                    <strong>Nota Importante:</strong> O valor da incerteza para a Resposta Analítica ({m_data['u_resposta_dinamica']:.4f}) reflete o desvio padrão calculado a partir das {m_data['n_reps_atual']} réplicas desta amostra.
                </p>
                """, unsafe_allow_html=True)
            st.markdown('</div>', unsafe_allow_html=True)

st.markdown("---")

# --- SEÇÃO: CAPACIDADE DE DECISÃO (CCα) ---
st.header("🛡️ Capacidade de Decisão (CCα) e Limites de Conformidade")

ativar_cca = st.checkbox("Calcular Capacidade de Decisão (CCα) para Avaliação de Risco", key="chk_ativar_cca")

if ativar_cca:
    if not st.session_state.mapa_incertezas:
        st.warning("⚠️ Execute primeiro a 'Avaliação de Incerteza' acima para alimentar este módulo.")
    else:
        amostras_disponiveis = list(st.session_state.mapa_incertezas.keys())
        
        c_cca1, c_cca2 = st.columns(2)
        with c_cca1:
            amostras_alvo = st.multiselect("Selecione as amostras para o cálculo de CCα:", options=amostras_disponiveis, default=amostras_disponiveis, key="cca_amostras_alvo")
        with c_cca2:
            jurisdicao = st.selectbox("Jurisdição / Órgão Regulador:", ["UE (União Europeia)", "Anvisa", "MAPA", "Codex Alimentarius", "EUA (FDA/EPA)"], key="cca_jurisdicao")
        
        c_cca3, c_cca4 = st.columns(2)
        with c_cca3:
            tipo_substancia = st.selectbox("Classificação do Analito:", ["Substância Autorizada (Possui LMR)", "Substância Proibida / Sem Autorização (Limite Zero / MRPL)"], key="cca_tipo_substancia")
        with c_cca4:
            blueprint = ["Incerteza / Diretriz UE 2021/808", "ISO 11843 (Base estatística da Curva)", "Validação no LMR (Dados Históricos de Repetibilidade)"]
            abordagem = st.selectbox("Abordagem Normativa / Metodologia:", blueprint, key="cca_abordagem")
            
        if jurisdicao == "UE (União Europeia)" and tipo_substancia == "Substância Proibida / Sem Autorização (Limite Zero / MRPL)":
            alfa_padrao = 1
            st.info("ℹ️ **Adaptação Automática:** Para substâncias proibidas na UE (Regulamento 2021/808), o risco de falso positivo é fixado em **α = 1%**.")
        else:
            alfa_padrao = 5
            st.info(f"ℹ️ **Adaptação Automática:** Para as diretrizes da {jurisdicao}, o risco monocaudal aceito para falsos positivos é de **α = 5%**.")

        lmr_val = st.number_input("Insira o Limite Máximo de Resíduo (LMR) ou Limite de Ação estabelecido (mg/Kg):", min_value=0.001, max_value=10.0, value=0.050, step=0.005, format="%.3f", key="cca_lmr_val")

        sr_historico = 0.0
        if abordagem == "Validação no LMR (Dados Históricos de Repetibilidade)":
            sr_historico = st.number_input("Insira o Desvio Padrão de Repetibilidade Histórica (s_R) obtido no LMR na validação (mg/Kg):", min_value=0.0001, max_value=1.0, value=0.0040, format="%.4f", key="cca_sr_historico")

        if st.button("⚖️ Executar Diagnóstico de Conformidade CCα"):
            if not amostras_alvo:
                st.error("⚠️ Selecione ao menos uma amostra no campo acima para gerar o parecer de CCα.")
            else:
                z_alfa = 2.326 if alfa_padrao == 1 else 1.645
                mapa_cca_resultados = {}
                
                for am_nome in amostras_alvo:
                    dados_am = st.session_state.mapa_incertezas[am_nome]
                    conc_medida = dados_am["conc"]
                    uc_amostra = dados_am["uc"]
                    
                    if abordagem == "Incerteza / Diretriz UE 2021/808":
                        if tipo_substancia == "Substância Proibida / Sem Autorização (Limite Zero / MRPL)":
                            cca_limite = 0.0 + z_alfa * uc_amostra
                        else:
                            cca_limite = lmr_val + z_alfa * uc_amostra
                        texto_limiar = f"CCα = {cca_limite:.3f} mg/Kg"
                            
                    elif abordagem == "ISO 11843 (Base estatística da Curva)":
                        if st.session_state.reg_results:
                            res_c = st.session_state.reg_results
                            if np.isclose(res_c["a"], 0):
                                st.error("⚠️ Não foi possível aplicar ISO 11843: a inclinação da curva está zerada ou muito próxima de zero.")
                                st.stop()
                            s_conc = res_c["epadyx"] / res_c["a"]
                            if tipo_substancia == "Substância Proibida / Sem Autorização (Limite Zero / MRPL)":
                                cca_limite = 0.0 + z_alfa * s_conc
                            else:
                                cca_limite = lmr_val + z_alfa * s_conc
                            texto_limiar = f"CCα = {cca_limite:.3f} mg/Kg"
                        else:
                            cca_limite = lmr_val
                            texto_limiar = "Curva não encontrada."
                    else:
                        if tipo_substancia == "Substância Proibida / Sem Autorização (Limite Zero / MRPL)":
                            cca_limite = 0.0 + z_alfa * sr_historico
                        else:
                            cca_limite = lmr_val + z_alfa * sr_historico
                        texto_limiar = f"CCα = {cca_limite:.3f} mg/Kg"

                    if conc_medida > cca_limite:
                        status_classe = "diag-warning"
                        status_msg = f"🚨 **NÃO CONFORME:** O valor medido ({conc_medida:.3f} mg/Kg) ultrapassa o limiar estatístico de decisão {texto_limiar}."
                        status_limpo = "NÃO CONFORME"
                    elif conc_medida > lmr_val:
                        status_classe = "diag-warning"
                        status_msg = f"⚠️ **ZONA DE INSEGURANÇA:** O valor medido ({conc_medida:.3f} mg/Kg) é superior ao LMR ({lmr_val:.3f} mg/Kg), mas abaixo da capacidade de decisão do método ({cca_limite:.3f} mg/Kg)."
                        status_limpo = "ZONA DE ALERTA"
                    else:
                        status_classe = "diag-good"
                        status_msg = f"✔️ **CONFORME:** O teor quantificado ({conc_medida:.3f} mg/Kg) está contido dentro dos limites legais de tolerância ({lmr_val:.3f} mg/Kg)."
                        status_limpo = "CONFORME"

                    mapa_cca_resultados[am_nome] = {
                        "status_classe": status_classe,
                        "status_msg": status_msg,
                        "status_limpo": status_limpo,
                        "cca_limite": cca_limite,
                        "lmr": lmr_val,
                        "abordagem": abordagem,
                        "alfa": alfa_padrao,
                        "texto_limiar": texto_limiar
                    }
                st.session_state.dados_cca_report = mapa_cca_resultados

if ativar_cca and st.session_state.dados_cca_report:
    st.markdown("### 📋 Resultados e Parecer de Capacidade de Decisão")
    for am_nome, res_cca in st.session_state.dados_cca_report.items():
        st.markdown(f"""
        <div class="res-box" style="border-left-color: #00d4ff;">
            <span style="font-size:15px; font-weight:bold; color:#0a2540;">Avaliação Legal de Risco — {am_nome}</span><br>
            <div style="font-size:12px; margin-top:2px; margin-bottom:8px; color:#555;">
                <b>Metodologia:</b> {res_cca['abordagem']} | <b>Risco Adotado (α):</b> {res_cca['alfa']}% | <b>LMR:</b> {res_cca['lmr']:.3f} mg/Kg
            </div>
            <div class="diag-box {res_cca['status_classe']}" style="margin-top:5px;">
                {res_cca['status_msg']}
            </div>
        </div>
        """, unsafe_allow_html=True)

st.markdown("---")

# --- SEÇÃO 4: EXPORTAÇÃO EXECUTIVA ---
if st.session_state.resultados_report:
    st.subheader("📥 Geração do Relatório Executivo")
    
    data_hora_sistemica = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    c_rep1, c_rep2 = st.columns(2)
    
    dados_export_cca = st.session_state.dados_cca_report if (ativar_cca and st.session_state.dados_cca_report) else None
    
    with c_rep1:
        excel_data = gerar_excel(st.session_state.resultados_report, analista_selecionado, data_hora_sistemica, LOQ_METODO, dados_cca=dados_export_cca)
        st.download_button(
            label="📊 Exportar Relatório Oficial para Excel",
            data=excel_data,
            file_name=f"Avaliacao_Incerteza_{composto.replace(' ', '_')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            width='stretch'
        )
        
    with c_rep2:
        pdf_data = gerar_pdf(st.session_state.resultados_report, analista_selecionado, data_hora_sistemica, dados_cca=dados_export_cca)
        st.download_button(
            label="📄 Exportar Relatório Oficial para PDF",
            data=pdf_data,
            file_name=f"Avaliacao_Incerteza_{composto.replace(' ', '_')}.pdf",
            mime="application/pdf",
            width='stretch'
        )
